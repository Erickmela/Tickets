"""
Utilidades para exportación de reportes
Responsabilidad: Generar archivos PDF y Excel
"""
from io import BytesIO
from datetime import datetime


class ExportadorPDF:
    """
    Clase para generar reportes en PDF
    """
    
    @staticmethod
    def generar_pdf(tipo_reporte, datos, filtros):
        """
        Generar archivo PDF del reporte
        
        TODO: Implementar generación de PDF con ReportLab o WeasyPrint
        Por ahora retorna un PDF básico de ejemplo
        """
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            elements = []
            
            # Estilos
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#1f2937'),
                alignment=TA_CENTER,
                spaceAfter=30,
            )
            
            # Título
            titulo_map = {
                'ventas': 'Reporte de Ventas',
                'validaciones': 'Reporte de Validaciones',
                'eventos': 'Reporte de Eventos',
                'personal': 'Reporte de Personal'
            }
            titulo = Paragraph(titulo_map.get(tipo_reporte, 'Reporte'), title_style)
            elements.append(titulo)
            
            # Fecha de generación
            fecha_generacion = Paragraph(
                f"<b>Fecha de generación:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                styles['Normal']
            )
            elements.append(fecha_generacion)
            elements.append(Spacer(1, 0.3 * inch))
            
            # Filtros aplicados
            if filtros.get('fecha_inicio') or filtros.get('fecha_fin'):
                filtros_text = "<b>Filtros aplicados:</b><br/>"
                if filtros.get('fecha_inicio'):
                    filtros_text += f"Desde: {filtros['fecha_inicio']}<br/>"
                if filtros.get('fecha_fin'):
                    filtros_text += f"Hasta: {filtros['fecha_fin']}<br/>"
                
                filtros_para = Paragraph(filtros_text, styles['Normal'])
                elements.append(filtros_para)
                elements.append(Spacer(1, 0.3 * inch))
            
            # Resumen (si existe)
            if 'resumen' in datos:
                resumen = datos['resumen']
                resumen_data = [
                    ['Métrica', 'Valor'],
                    ['Total Ventas', str(resumen.get('total_ventas', 0))],
                    ['Total Tickets', str(resumen.get('total_tickets', 0))],
                    ['Total Validados', str(resumen.get('total_validados', 0))],
                    ['Ingresos Totales', f"S/ {resumen.get('ingresos_totales', 0):.2f}"],
                ]
                
                resumen_table = Table(resumen_data, colWidths=[3*inch, 2*inch])
                resumen_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                elements.append(Paragraph("<b>Resumen General</b>", styles['Heading2']))
                elements.append(Spacer(1, 0.2 * inch))
                elements.append(resumen_table)
                elements.append(Spacer(1, 0.4 * inch))
            
            # Detalles
            if 'detalles' in datos and datos['detalles']:
                elements.append(Paragraph("<b>Detalle</b>", styles['Heading2']))
                elements.append(Spacer(1, 0.2 * inch))
                
                # Limitar a primeros 50 registros para el PDF
                detalles = datos['detalles'][:50]
                
                # Construir tabla según tipo de reporte
                if tipo_reporte == 'ventas':
                    detalle_data = [['ID', 'Fecha', 'Vendedor', 'Cliente', 'Total']]
                    for venta in detalles:
                        detalle_data.append([
                            str(venta.get('id', '')),
                            venta.get('fecha', '')[:10],
                            venta.get('vendedor', ''),
                            venta.get('cliente', '')[:20],
                            f"S/ {venta.get('total', 0):.2f}"
                        ])
                elif tipo_reporte == 'eventos':
                    detalle_data = [['ID', 'Evento', 'Fecha', 'Vendidos', 'Disponibles', 'Ocupación']]
                    for evento in detalles:
                        detalle_data.append([
                            str(evento.get('id', '')),
                            evento.get('nombre', '')[:30],
                            evento.get('fecha', '')[:10],
                            str(evento.get('vendidos', 0)),
                            str(evento.get('disponibles', 0)),
                            f"{evento.get('ocupacion', 0)}%"
                        ])
                else:
                    # Generic table para otros tipos
                    detalle_data = [['Información del Reporte']]
                    detalle_data.append([f'Total de registros: {len(detalles)}'])
                
                detalle_table = Table(detalle_data)
                detalle_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                ]))
                
                elements.append(detalle_table)
            
            # Generar PDF
            doc.build(elements)
            pdf_content = buffer.getvalue()
            buffer.close()
            
            return pdf_content
        
        except ImportError:
            # Si no está instalado ReportLab, retornar PDF básico
            return b'%PDF-1.4\nReporte generado. Instalar reportlab para PDFs completos.'
        except Exception as e:
            raise Exception(f"Error al generar PDF: {str(e)}")


class ExportadorExcel:
    """
    Clase para generar reportes en Excel
    """
    
    @staticmethod
    def generar_excel(tipo_reporte, datos, filtros):
        """
        Generar archivo Excel del reporte
        
        TODO: Implementar generación de Excel con openpyxl o xlsxwriter
        Por ahora retorna un Excel básico de ejemplo
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            
            # Título del reporte
            titulo_map = {
                'ventas': 'Reporte de Ventas',
                'validaciones': 'Reporte de Validaciones',
                'eventos': 'Reporte de Eventos',
                'personal': 'Reporte de Personal'
            }
            ws.title = titulo_map.get(tipo_reporte, 'Reporte')[:31]  # Max 31 chars
            
            # Estilos
            header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            # Título
            ws['A1'] = titulo_map.get(tipo_reporte, 'Reporte')
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:E1')
            
            # Fecha de generación
            ws['A2'] = 'Fecha de generación:'
            ws['B2'] = datetime.now().strftime('%d/%m/%Y %H:%M')
            
            # Filtros
            row = 4
            if filtros.get('fecha_inicio'):
                ws[f'A{row}'] = 'Fecha Inicio:'
                ws[f'B{row}'] = str(filtros['fecha_inicio'])
                row += 1
            if filtros.get('fecha_fin'):
                ws[f'A{row}'] = 'Fecha Fin:'
                ws[f'B{row}'] = str(filtros['fecha_fin'])
                row += 1
            
            row += 1
            
            # Resumen
            if 'resumen' in datos:
                resumen = datos['resumen']
                ws[f'A{row}'] = 'RESUMEN GENERAL'
                ws[f'A{row}'].font = Font(bold=True, size=12)
                row += 1
                
                ws[f'A{row}'] = 'Total Ventas:'
                ws[f'B{row}'] = resumen.get('total_ventas', 0)
                row += 1
                
                ws[f'A{row}'] = 'Total Tickets:'
                ws[f'B{row}'] = resumen.get('total_tickets', 0)
                row += 1
                
                ws[f'A{row}'] = 'Total Validados:'
                ws[f'B{row}'] = resumen.get('total_validados', 0)
                row += 1
                
                ws[f'A{row}'] = 'Ingresos Totales:'
                ws[f'B{row}'] = f"S/ {resumen.get('ingresos_totales', 0):.2f}"
                row += 2
            
            # Detalles
            if 'detalles' in datos and datos['detalles']:
                ws[f'A{row}'] = 'DETALLE'
                ws[f'A{row}'].font = Font(bold=True, size=12)
                row += 1
                
                detalles = datos['detalles']
                
                # Headers según tipo
                if tipo_reporte == 'ventas':
                    headers = ['ID', 'Fecha', 'Vendedor', 'Cliente', 'Tickets', 'Total']
                    ws.append(headers)
                    
                    for venta in detalles:
                        ws.append([
                            venta.get('id', ''),
                            venta.get('fecha', ''),
                            venta.get('vendedor', ''),
                            venta.get('cliente', ''),
                            venta.get('tickets', 0),
                            venta.get('total', 0)
                        ])
                
                elif tipo_reporte == 'eventos':
                    headers = ['ID', 'Evento', 'Fecha', 'Vendidos', 'Disponibles', 'Ocupación %']
                    ws.append(headers)
                    
                    for evento in detalles:
                        ws.append([
                            evento.get('id', ''),
                            evento.get('nombre', ''),
                            evento.get('fecha', ''),
                            evento.get('vendidos', 0),
                            evento.get('disponibles', 0),
                            evento.get('ocupacion', 0)
                        ])
                
                elif tipo_reporte == 'personal':
                    headers = ['ID', 'Nombre', 'Rol', 'Ventas/Validaciones', 'Total']
                    ws.append(headers)
                    
                    for persona in detalles:
                        ws.append([
                            persona.get('id', ''),
                            persona.get('nombre', ''),
                            persona.get('rol', ''),
                            persona.get('ventas', 0),
                            persona.get('total', 0)
                        ])
                
                # Aplicar estilo a headers
                for cell in ws[row + 1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
            
            # Ajustar anchos de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Guardar en buffer
            buffer = BytesIO()
            wb.save(buffer)
            excel_content = buffer.getvalue()
            buffer.close()
            
            return excel_content
        
        except ImportError:
            # Si no está instalado openpyxl, retornar archivo básico
            raise Exception("Debe instalar openpyxl para generar archivos Excel")
        except Exception as e:
            raise Exception(f"Error al generar Excel: {str(e)}")
